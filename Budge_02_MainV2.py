# -*- coding: utf-8 -*-
"""
Created on Tue Jan 12 10:53:45 2021

@author: Andrés Cueto Estrada
@Programa: Main V2, Lector de catálogo - Budge_02_MainV2

@Descripción: n vista que no pude importar clases y etc... haré mi desmadre aquí mismo 
    en un solo script
"""






# ----------------------------------------- MÓDULO 1
"""
@Descripción: En esta sección se pueden observar todos los datos que el usuario deberá 
    ingresar para que el programa pueda trabajar, tales como:
        - Ubicación de las carpetas:
            - Carpeta en donde se ubican las Particiones. Variable: directorio_particiones
            - Carpeta en donde se ubica el catálogo. Variables: directorio_catálogo
            - Columnas en donde cada tipo de dato se encuentra
            - Renglones de inicio en cada uno de los documentos
            - Obra en la que se trabaja
"""

print('...Leyendo datos ingresados...')
#---------------------------   Datos ingresados para trabajar  ---------------------------
#---------------------------   M1.1 Carpetas de Origen   ---------------------------
directorio_catálogo = r'C:\Users\ADMIN\Google Drive\LIVERPOOL LA PERLA\4.- ESTIMACIONES\1. Catalogo y contratos\2. Catalogo\Catalogo Liverpool La Perla Claves Margaleff.xlsx'  
directorio_particiones = r'C:\Users\ADMIN\Google Drive\LIVERPOOL LA PERLA\12.- CUANTIFICACION PROYECTO\PARTIDAS'
directorio_destino = r'C:\Users\ADMIN\Google Drive\LIVERPOOL LA PERLA\28.- AUTOMATIZACIONES\2 Budge - Buscador de cantidades totales de materiales\1 Reportes'  # Destino del reporte


#---------------------------   M1.2.1 Ubicación de celdas de información CATÁLOGO   ---------------------------
# NOTA: Ingrese las columnas en MAYÚSCULAS
catalogo_nombre_hoja = 'CAT MARGALEFF'
catalogo_columna_matricula = 'B'
catalogo_columna_nombre = 'C'
catalogo_columna_unidad = 'D'
catalogo_columna_cantidad = 'E'
catalogo_columna_pu = 'F'
catalogo_renglon_inicio = 4  # Primer renglón en el que aparecen datos de forma contínua


#---------------------------   M1.2.2 Ubicación de celdas de información PARTICIONES   ---------------------------
particiones_nombre_hoja = 'RES GRAL'
particiones_columna_matricula = 'B'
particiones_columna_nombre = 'C'
particiones_columna_unidad = 'D'
particiones_columna_cantidad = 'E'
particiones_renglon_inicio = 17  # Primer renglón en el que aparecen datos de forma contínua





#---------------------------   M1.3 NOMBRE DEL REPORTE FINAL   ---------------------------
nombre_reporte = 'Cantidades totales de materiales'


















#---------------------------   ¡¡¡¡¡¡¡¡¡¡¡¡¡ATENCIÓN!!!!!!!!!!!!!   ---------------------------
# De aquí en adelante, el código no deberá ser editado sin conocimiento del lenguaje,
#   pues cualquier cambio podría afectar en el desempeño del programa y/o arrojar datos
#   erroneos.




















print('...Definiendo clases...')
class Documento:
    nombre_hoja = ''
    columna_matricula = ''
    columna_nombre = ''
    columna_unidad = ''
    columna_pu = ''
    columna_cantidad = ''
    columna_cantidad_2 = ''
    renglon_inicio = 0
    
    
    celda_matricula_actual = ''
    celda_nombre_actual = ''
    celda_unidad_actual = ''
    celda_pu_actual = ''
    celda_cantidad_actual = ''    
    celda_cantidad_2_actual = ''    
    renglon_actual = 0
    
    def siguiente_renglon_ (self):
        # Aquí algo está pasando que nunca se ejecuta
        if self.renglon_actual == 0:
            self.renglon_actual = self.renglon_inicio
            self.generando_celdas_actuales()
        else:
             self.renglon_actual += 1
             self.generando_celdas_actuales()
    
    def generando_celdas_actuales (self):
        renglon_actual_str = str(self.renglon_actual)
        self.celda_matricula_actual = self.columna_matricula + renglon_actual_str
        self.celda_nombre_actual = self.columna_nombre + renglon_actual_str
        self.celda_unidad_actual = self.columna_unidad + renglon_actual_str
        self.celda_cantidad_actual = self.columna_cantidad + renglon_actual_str
        self.celda_cantidad_2_actual = self.columna_cantidad_2 + renglon_actual_str
        self.celda_pu_actual = self.columna_pu + renglon_actual_str
        
        
    
catalogo = Documento()
catalogo.nombre_hoja = catalogo_nombre_hoja
catalogo.columna_matricula = catalogo_columna_matricula
catalogo.columna_nombre = catalogo_columna_nombre
catalogo.columna_unidad = catalogo_columna_unidad  
catalogo.columna_cantidad = catalogo_columna_cantidad
catalogo.columna_pu = catalogo_columna_pu 
catalogo.renglon_inicio = catalogo_renglon_inicio


particiones = Documento()
particiones.nombre_hoja = particiones_nombre_hoja
particiones.columna_matricula = particiones_columna_matricula
particiones.columna_nombre = particiones_columna_nombre
particiones.columna_unidad = particiones_columna_unidad  
particiones.columna_cantidad = particiones_columna_cantidad
particiones.renglon_inicio = particiones_renglon_inicio



















# ----------------------------------------- MÓDULO 2
"""
@Descripción: En este script se pueden observar la declaración del Objeto tipo Material 
    este tipo de objeto se definirá según su:
        - Material (nombre)
        - Matrícula
        - Precio Unitario
        - Cantidad presupuestada por el proveedor
        - Cantidad Contabilizada
    también conteiene algunos métodos que el mismo objeto requerirá, como la suma de los
    siguientes valores que se le ingresen.
"""
class Material:
    def __init__(self, matricula, nombre, unidad, precio_unitario, cantidad_proveedor, cantidad_contabilizada):
        self.matricula = matricula
        self.nombre = nombre
        self.unidad = unidad
        self.precio_unitario = precio_unitario
        self.cantidad_proveedor = cantidad_proveedor
        self.cantidad_contabilizada = cantidad_contabilizada
        
    matricula = ''
    nombre = ''
    unidad = ''
    precio_unitario = ''
    cantidad_proveedor = ''
    cantidad_contabilizada = ''
    
    
    
    
    


    















print('...Leyendo directorio de particiones...')
# ----------------------------------------- MÓDULO 3
"""
@Descripción: Este script ejecuta los módulos M1 y M2 para poder obtener sus datos y 
    trabajar con ellos con la finalidad de leer las carpetas cargadas en M1
"""
import os
def lector_archivos_desde_carpetas (directorio, extension_permisible):
    lista_archivos = os.listdir(directorio)
    
    archivos_xlsx = []
    for elemento in lista_archivos:
        # Condiciones
        condicion_extension = elemento.find(extension_permisible) > 0
        condicion_abierto_drive = elemento.find('~') == 0
        # Si en un futuro se agregan más condiciones, una forma elegante y simple de leerlas será un ciclo for
        # así que se generará una lista:
        condiciones= []
        condiciones.append(condicion_extension)
        condiciones.append(condicion_abierto_drive)
        
        # Obtenemos el resultado de la condicion final
        condicion_final = True
        for condicion in condiciones:
            condicion_final = condicion_final and condicion
        
        # Si está abierto e drive
        if condicion_abierto_drive:
            print('El archivo {} no se leerá ya que se encuentra actualmente abierto por Google Drive'.format(elemento))
        
        
        # Si todas las condiciones se cumplen, se guardará en la lista de archivos
        if condicion_final:
            archivos_xlsx.append(elemento)
        
            
    
    lista_archivos = archivos_xlsx
    
    
    
    return lista_archivos

# Leemos el directorio de las particiones y obtenemos la lista de archivos de particiones
lista_archivos = lector_archivos_desde_carpetas (directorio_particiones, '.xlsx')















print('...Leyendo catálogo...')
# ----------------------------------------- MÓDULO 4 Lector de Catálogo
# ----------------------------------------- Abriendo documento
""" 
Aquí guardaremos los datos leidos del Excel, por lo que deberemos primero generar 
        una iteración entre renglones y sacar cada uno de los datos
        - Generar iteración con for
        - Generar lectura de una celda
        - Generar comprobación de información
        - Guardar en lista
"""

import openpyxl as oppy
from openpyxl import Workbook 

wb = oppy.load_workbook(directorio_catálogo)
ws = wb[catalogo.nombre_hoja]


espacios_en_blanco = 0
diccionario_materiales_catalogo = {}

while espacios_en_blanco < 10:
    # Definir celda actual en la que se trabajará
    catalogo.siguiente_renglon_()
    
    # Pasamos las celdas de string a tipo celda
    celda_matricula = ws[catalogo.celda_matricula_actual]
    celda_nombre = ws[catalogo.celda_nombre_actual]
    celda_unidad = ws[catalogo.celda_unidad_actual]
    celda_cantidad = ws[catalogo.celda_cantidad_actual]
    celda_pu = ws[catalogo.celda_pu_actual]
    
    # Vemos que celdas están vacías
    celda_matricula_vacia = celda_matricula.value is None  # TRUE Si está vacía
    celda_nombre_vacia = celda_nombre.value is None  # TRUE Si está vacía
    celda_unidad_vacia = celda_unidad.value is None  # TRUE Si está vacía
    celda_cantidad_vacia = celda_cantidad.value is None  # TRUE Si está vacía
    celda_pu_vacia = celda_pu.value is None  # TRUE Si está vacía
    
    
    #print(celda_matricula)
    
    
    alguna_celda_vacia = celda_matricula_vacia or celda_nombre_vacia or celda_unidad_vacia or celda_cantidad_vacia or celda_pu_vacia
    
    
    
    if alguna_celda_vacia:
        #Si no hay datos completos en ese renglon, no se hará nada, mas que aumentar el contador de espacios en blanco
        espacios_en_blanco += 1        
    else:
        #Pero, si están todos los valores, entonces se procede a guardarlos
        espacios_en_blanco = 0
        # Aquí se guardarán los datos
        #  diccionario_materiales_catalogo
        matricula = str(celda_matricula.value)
        
        if matricula in diccionario_materiales_catalogo.keys():
            # Si el valor ya está ahí, pues que raro la verdad, debe estar repetido
            #   pero bueno, se suman las cantidades para obtener el total
            
            
            # Sumamos la cantidad anterior con la actual, convirtiéndola en int, primero
            primer_cantidad = float(diccionario_materiales_catalogo[matricula].cantidad_proveedor)
            segunda_cantidad = float(celda_cantidad.value)
            
            diccionario_materiales_catalogo[matricula].cantidad_proveedor = str( primer_cantidad + segunda_cantidad )
            
            #print()
            #print('La matrícula {} se ha encontrado repetida en la celda {}'.format(celda_matricula.value, celda_matricula))
            #print('La cantidad total hasta el momento es: {} '.format(diccionario_materiales_catalogo[celda_matricula.value].cantidad_proveedor))
        else:
            # En otro caso, lo agregamos al diccionario
            
            # A partir de las celdas actuales, obtenemos los valores en dichas celdas
            #   y se lo otorgamos a su variable local correspondiente
            matricula = str(celda_matricula.value)
            nombre = celda_nombre.value
            unidad = celda_unidad.value
            precio_unitario = celda_pu.value
            cantidad_proveedor = celda_cantidad.value
            
            # Al diccionario le agregamos la matrícula como key y el objeto materia como value
            #   usamos el constructor para, de una vez, agregar los valores del objeto "material"
            diccionario_materiales_catalogo[ matricula ] = Material(matricula, nombre, unidad, precio_unitario, cantidad_proveedor, None)
            #aux = diccionario_materiales_catalogo[celda_matricula.value]
            # HASTA ESTE MOMENTO, LAS CANTIDADES DEL PROVEEDOR SE SUMAN Y SE TIENE UN TOTAL EN EL DICCIOANRIO
            #   POR LO TANTO LA PRIMER PARTE DEL PROGRAMA ESTÁ TERMINADA, YA OBTUVIMOS LAS MATRÍCULAS A PARTIR DEL CATÁLOGO
            #   FALTA HACER LO MISMO CON LAS PARTIDAS
        
            
    
        
        














print('...Leyendo particiones...')
# ----------------------------------------- MÓDULO 5 Lector de Particiones
""" 
Aquí guardaremos los datos leidos a partir de cada uno de los Excel dentro de la carpeta de particiones, 
    por lo que primero deberemos generar una iteración entre los documentos.
        - Iteración entre documentos
        - Iteración entre cada renglón del documento
        - Validar información y guardarla
"""

# Al final se deberá obtener los siguientes diccionarios:
diccionario_materiales_particiones = {}
diccionario_materiales_sin_matricula = {}
# Iterando entre cada uno de los archivos de la carpeta de particiones
for archivo in lista_archivos:
    # Generando su dirección completa para poderlo leer
    directorio_particion_n = directorio_particiones + '\\' + archivo
    
    # Abriendo el woorkbook actual
    wb = oppy.load_workbook(directorio_particion_n , data_only = True)
    ws = wb[particiones.nombre_hoja]
    
    # Inicializamos nuestra variable
    espacios_en_blanco = 0
    particiones.renglon_actual = particiones.renglon_inicio    
    
    # Empezamos el ciclo while para obtener datos
    while espacios_en_blanco < 10:
        # Definir celda actual en la que se trabajará
        particiones.siguiente_renglon_()
        
        # Pasamos las celdas de string a tipo celda
        celda_matricula = ws[particiones.celda_matricula_actual]
        celda_nombre = ws[particiones.celda_nombre_actual]
        celda_unidad = ws[particiones.celda_unidad_actual]
        celda_cantidad = ws[particiones.celda_cantidad_actual]
        #celda_pu = ws[particiones.celda_pu_actual]
        
        # Vemos que celdas están vacías
        celda_matricula_vacia = celda_matricula.value is None  # TRUE Si está vacía
        celda_nombre_vacia = celda_nombre.value is None  # TRUE Si está vacía
        celda_unidad_vacia = celda_unidad.value is None  # TRUE Si está vacía
        celda_cantidad_vacia = celda_cantidad.value is None  # TRUE Si está vacía
        #celda_pu_vacia = celda_pu.value is None  # TRUE Si está vacía
        
            
        # Generamos un condicionar que de TRUE si todas las celdas especificadas cumplen con el requisito de estar llenas
        alguna_celda_vacia =  celda_nombre_vacia or celda_unidad_vacia or celda_cantidad_vacia
        
        # La condición especial establece que si se tienen todos los datos excepto la matrícula, 
        #   se registrará en un diccionario especial. Para esto se tendrá lo siguiente:
        #   alguna celda vacía deberá ser FALSO, indicando que todas tienen datos
        #   celda_matricula_vacía deberá ser TRUE, porque si deberá estar vacío
        #   por lo tanto la condición será: (not alguna_celda_vacia) and celda_matricula_vacía
        condicion_especial = ( celda_matricula_vacia and (not alguna_celda_vacia) )
        
        
        
        
        if alguna_celda_vacia:
            #Si no hay datos completos en ese renglon, no se hará nada, mas que aumentar el contador de espacios en blanco
            espacios_en_blanco += 1
        elif condicion_especial:
            # Si se tiene todo excepto la matrícula se deberá  guardar en un diccionario especial
            espacios_en_blanco = 0
            nombre = str(celda_nombre.value)
            
            if nombre in diccionario_materiales_sin_matricula:
                # Si el nombre ya está registrado, sumaremos los totales que hemos encontrado
                aux = float(diccionario_materiales_sin_matricula[nombre].cantidad_contabilizada)
                diccionario_materiales_sin_matricula[nombre].cantidad_contabilizada = str( aux + float(celda_cantidad.value))
            else:
                # Si el nombre no está registrado, lo registraremos
                unidad = celda_unidad.value
                precio_unitario = celda_pu.value
                cantidad_contabilizada = celda_cantidad.value
                
                diccionario_materiales_sin_matricula[nombre] = Material( None , nombre, unidad, None, None , cantidad_contabilizada)
        
        else:
            #Pero, si están todos los valores, entonces se procede a guardarlos
            espacios_en_blanco = 0
            # En esta sección se guardarán los datos
            
            matricula = str(celda_matricula.value)
            if matricula in diccionario_materiales_particiones.keys():
                # Si el valor ya está ahí debe estar repetido y las cantidades se suman para obtener el total
                
                
                # Sumamos la cantidad anterior con la actual, convirtiéndola en int, primero
                #print('cantidad contabilizada: ',diccionario_materiales_particiones[celda_matricula.value].cantidad_contabilizada)
                primer_cantidad = float(diccionario_materiales_particiones[matricula].cantidad_contabilizada)
                segunda_cantidad = float(celda_cantidad.value)
                
                
                diccionario_materiales_particiones[matricula].cantidad_contabilizada = str( primer_cantidad + segunda_cantidad )
                
            else:
                # En otro caso, lo agregamos al diccionario
                
                # A partir de las celdas actuales, obtenemos los valores en dichas celdas
                #   y se lo otorgamos a su variable local correspondiente
                matricula = str(celda_matricula.value)
                nombre = celda_nombre.value
                unidad = celda_unidad.value
                precio_unitario = celda_pu.value
                cantidad_contabilizada = celda_cantidad.value
                
                # Al diccionario le agregamos la matrícula como key y el objeto "Material" como value
                #   usamos el constructor para, de una vez, agregar los valores del objeto "material"
                diccionario_materiales_particiones[ matricula ] = Material(matricula, nombre, unidad, None, None , cantidad_contabilizada)
                #aux = diccionario_materiales_catalogo[celda_matricula.value]




















print('...Procesando información...')
# ----------------------------------------- MÓDULO 6 Procesador de datos
""" 
Aquí procesaremos la información de los 3 diccionarios actuales:
    - diccionario_materiales_catalogo
    - diccionario_materiales_particiones
    - diccionario_materiales_sin_matricula
    
    Crearemos un nuevo diccionario con matrículas de materiales 
    1.- Crear nuevo diccionario
    2.- Copiar el diccionario de particiones (porque es el más aplio)
    3.- Checar si ya tiene los materiales del catálogo
        3.1.- Si no, agregarlos con sus datos originales del primer diccionario
        3.2.- Si si, solo agregar CONTEO DE PROVEEDOR
    4.- Agregar los elementos del diccionario que carece de matrículas
        4.3.- Agregaremos su matrícula como elemento None
    
    5.- NOTA: Para el siguiente módulo, si se encuentra un eleento None, según el condicional
    que generemos, será el color que se le de al elemento
"""                
           
# Creando el nuevo diccionario que tendrá todos los valores     
diccionario_materiales_completo = {}
# Asignando el diccionario más completo
diccionario_materiales_completo = diccionario_materiales_particiones.copy()

# Generando ciclo for para agregar el resto de keys
for material in diccionario_materiales_catalogo:
    
    # Checando si el material ya está en el diccionario
    if material in diccionario_materiales_completo :
        # Si ya existe el material solo habrá que colocar el valor conteo en catáloco
        diccionario_materiales_completo[material].cantidad_proveedor = diccionario_materiales_catalogo[material].cantidad_proveedor
        diccionario_materiales_completo[material].precio_unitario = diccionario_materiales_catalogo[material].precio_unitario
    else:
        # Si aún no existe el material lo otorgaremos
        diccionario_materiales_completo[material] = diccionario_materiales_catalogo[material]


# Filtraremos el diccionario de materiales completo, pues hay residuos aún
#   estos residuos son de subtítulos o títulos, una de las formas de deshacernos de ellos es 
#   identificar las unidades que podría haber en los materiales y no dejar pasar ninguna otra
#   creo que sería más fácil hacerlo por la longitud de estas unidades
# O quizás el hecho de que no deberá contener espacios respecto a otras palabras
        


# Generando un for para poder evaluar la longitud del elemento que contiene
#   material.unidad, se genera un diccionario auxiliar
diccionario_auxiliar = {}
for elemento in diccionario_materiales_completo:
    # Cualquier unidad de un material no tendrá espacios entre palabras, porque será una sola
    # las unidades que hay no son demasiado largas, hay que estimar un largo
    
    material = diccionario_materiales_completo[elemento]
    # Removemos los espacios a las orillas
    unidad = material.unidad
    unidad = unidad.strip()
    
    # Empleando la primer condición, existencia de espacios:
    condicion_espacios = unidad.find(' ') > 0
    # Empleando la segunda condición, longitud de la cadena:
    condicion_longitud = len(unidad) > 15
    
        
    if condicion_espacios or condicion_longitud:
        # Si si tiene espacios o es muy larga, no se guardará el elemento
        pass
       # print('Se ha eliminado \t Material: {} \t Unidad: {}'.format(matricula, unidad))
    else:
         # Si no tiene espacios o es lo suficientemente corta, se acepta en el diccionario
         diccionario_auxiliar[elemento] = material
         
diccionario_materiales_completo.clear()
diccionario_materiales_compelto = diccionario_auxiliar.copy()

from datetime import datetime
fecha = datetime.now().strftime('%d-%m-%y')    
cantidad_documentos_leidos = str(len(lista_archivos))
cantidad_materiales = str( len( diccionario_materiales_compelto ) )
cantidad_materiales_con_matricula = str( len( diccionario_materiales_catalogo ) )


cantidad_materiales_sin_matricula = str( len(diccionario_materiales_compelto) - len( diccionario_materiales_catalogo ) )
if float(cantidad_materiales_sin_matricula) >= 0:
    # Esto significa que si hay materiales sin matrícula o que son 0
    pass
else:
    # Si el signo es negativo, significa que el catálogo tiene más materiales que el listado completo
    #   esto presenta una situación imprevista, ya que el completo está formato por el catálogo 
    #   más las particiones, así que sería un error
    cantidad_materiales_sin_matricula = 'Error línea 609'



















# ----------------------------------------- MÓDULO 7 Generador de reporte
print('...Generando Reporte...')
wb = Workbook()
ws = wb.active

#  7.1 Generar todas las cadenas que generaremos en el excel
print('..Generando textos...')
# Generando solo las cadenas
titulo_general = 'Desgloce de la cantidad total de materiales'.title()

fecha_label = 'Fecha y hora de corte'
documentos_label = 'Documentos leidos'
materiales_label = 'Materiales'
materiales_con_matricula_label = 'Materiales con matricula'
materiales_sin_matricula_label = 'Materiales sin matricula'

buscador_titulo = 'Buscador de materiales por matrícula'.title()
matricula_label = 'Matrícula'.title()
nombre_label = 'Nombre'.title()
unidad_label = 'Unidad'.title()
precio_unitario = 'Precio unitario'.title()
cantidad_proveedor = 'Cantidad proveedor'.title()
cantidad_contabilizada = 'Cantidad contabilizada'.title()

listado_titulo = 'Listado de materiales'.title()



#  7.2 Generar el formato que usaremos para cada celda
print('...Definiendo celdas...')
#  7.2.1 Definiendo las celdas y rangos que usaremos
#        Celdas de información
celda_titulo_general = 'A1'
celdas_titulo_general = 'A1:M1'

celda_fecha_label = 'A3'
celdas_fecha_label = 'A3:C3'
celda_fecha = 'D3'

celda_documentos_label = 'A4'
celdas_documentos_label = 'A4:C4'
celda_documentos = 'D4'

celda_materiales_label = 'A5'
celdas_materiales_label = 'A5:C5'
celda_materiales = 'D5'

celda_materiales_con_matricula_label = 'J3'
celdas_materiales_con_matricula_label = 'J3:L3'
celda_materiales_con_matricula = 'M3'

celda_materiales_sin_matricula_label = 'J4'
celdas_materiales_sin_matricula_label = 'J4:L4'
celda_materiales_sin_matricula = 'M4'


#       Celdas de buscador
celda_buscador_label = 'A8'
celdas_buscador_label = 'A8:M8'

celda_matricula_label = 'A9'
celdas_matricula_label = 'A9:B9'
celda_matricula = 'A10'
celdas_matricula = 'A10:B10'

celda_nombre_label = 'C9'
celdas_nombre_label = 'C9:F9'
celda_nombre = 'C10'
celdas_nombre = 'C10:F10'

celda_unidad_label = 'G9'
celda_unidad = 'G10'

celda_pu_label = 'H9'
celdas_pu_label = 'H9:I9'
celda_pu = 'H10'
celdas_pu = 'H10:I10'

celda_cantidad_proveedor_label = 'J9'
celdas_cantidad_proveedor_label = 'J9:K9'
celda_cantidad_proveedor = 'J10'
celdas_cantidad_proveedor = 'J10:K10'

celda_cantidad_contabilizada_label = 'L9'
celdas_cantidad_contabilizada_label = 'L9:M9'
celda_cantidad_contabilizada = 'L10'
celdas_cantidad_contabilizada = 'L10:M10'


#       Celdas de listado
celda_listado_de_materiales_label = 'A13'
celdas_listado_de_materiales_label = 'A13:M13'

celda_matricula_label_2 = 'A14'
celdas_matricula_label_2 = 'A14:B14'

celda_nombre_label_2 = 'C14'
celdas_nombre_label_2 = 'C14:F14'

celda_unidad_label_2 = 'G14'

celda_pu_label_2 = 'H14'
celdas_pu_label_2 = 'H14:I14'

celda_cantidad_proveedor_label_2 = 'J14'
celdas_cantidad_proveedor_label_2 = 'J14:K14'

celda_cantidad_contabilizada_label_2 = 'L14'
celdas_cantidad_contabilizada_label_2 = 'L14:M14'




#       Celdas iniciales para vaciar información
reporte_doc = Documento()
reporte_doc.columna_matricula = 'A'
reporte_doc.columna_nombre = 'C'
reporte_doc.columna_unidad = 'G'
reporte_doc.columna_pu = 'H'
reporte_doc.columna_cantidad = 'J'
reporte_doc.columna_cantidad_2 = 'L'
reporte_doc.renglon_inicio = 15





#  7.2.2 Combinando las celdas necesarias
print('...Combinando celdas...')
ws.merge_cells(celdas_titulo_general)
#       Celdas de datos
ws.merge_cells(celdas_fecha_label)
ws.merge_cells(celdas_documentos_label)
ws.merge_cells(celdas_materiales_label)
ws.merge_cells(celdas_materiales_con_matricula_label)
ws.merge_cells(celdas_materiales_sin_matricula_label)
#       Celdas del buscador
ws.merge_cells(celdas_buscador_label)
ws.merge_cells(celdas_matricula_label)
ws.merge_cells(celdas_matricula)
ws.merge_cells(celdas_nombre_label)
ws.merge_cells(celdas_nombre)
ws.merge_cells(celdas_pu_label)
ws.merge_cells(celdas_pu)
ws.merge_cells(celdas_cantidad_proveedor_label)
ws.merge_cells(celdas_cantidad_proveedor)
ws.merge_cells(celdas_cantidad_contabilizada_label)
ws.merge_cells(celdas_cantidad_contabilizada)
#       Celdas del listado
ws.merge_cells(celdas_listado_de_materiales_label)
ws.merge_cells(celdas_matricula_label_2)
ws.merge_cells(celdas_nombre_label_2)
ws.merge_cells(celdas_pu_label_2)
ws.merge_cells(celdas_cantidad_proveedor_label_2)
ws.merge_cells(celdas_cantidad_contabilizada_label_2)





#  7.2.3 Darle formato a las celdas (font, color, márgenes)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Función para rellenar los bordes
import re  # Para hacer split de cadenas
def bordes_en_combinacion(celda_izquierda_arriba , celda_derecha_abajo, configuracion_borde):
    """ 
    Ya que la generación de bordes tiene problemas para celdas combinadas por sí sola,
    esta función le da los bordes a las celdas combinadas ya sin ningún problema.
    - Recibe las celdas como strings"""
    # Estilos 
    
    # Obteniendo rangos
    celda_izquierda_arriba_split = re.split('(\d+)',celda_izquierda_arriba)
    celda_derecha_abajo_split = re.split('(\d+)',celda_derecha_abajo)
    distancia_row = abs(int(celda_izquierda_arriba_split[1]) - int(celda_derecha_abajo_split[1])) + 1
    distancia_columna = abs(ord(celda_izquierda_arriba_split[0]) - ord(celda_derecha_abajo_split[0])) 
    #print(distancia_columna)
    for renglon in range (distancia_row ):
        renglon_real = renglon + int(celda_izquierda_arriba_split[1])
        renglon_str = str(renglon_real)
        #print('trabajando en renglón: ' , renglon_str)
        for columna in range (distancia_columna + 1):
            columna_real = ord(celda_izquierda_arriba_split[0]) + columna
            columna_str = chr(columna_real)
            #print('trabajando en columna: ' , columna_str)
            celda = columna_str + renglon_str
            #print(celda)
            ws[celda].border = Border(top = configuracion_borde, left = configuracion_borde, right = configuracion_borde ,  bottom = configuracion_borde)    

# 7.2.3.1 Aquí se definirán los formatos: Side(bordes) , Alignment(alineaciones) , Fill(color de celda)  , Font(letra) 
# 7.2.3.1.1 Side(bordes)
borde = Side(border_style="thin", color="000000")

# 7.2.3.1.2 Alignment(alineaciones)
izquierda = Alignment( horizontal = 'left' , vertical = 'center' )
derecha = Alignment( horizontal = 'right' , vertical = 'center' )
centro = Alignment( horizontal = 'center' , vertical = 'center' )

# 7.2.3.1.3 Fill(color de celda)
negro = PatternFill("solid", fgColor='00000000')
gris = PatternFill("solid", fgColor='00D9D9D9')
naranja_fuerte = PatternFill("solid", fgColor='00FFC000')
naranja_debil = PatternFill("solid", fgColor='00FFE699')
azul_fuerte = PatternFill("solid", fgColor='0000B0F0')
azul_debil = PatternFill("solid", fgColor='00D9E1F2')
rojo = PatternFill("solid", fgColor='00FF0000')

# 7.2.3.1.4 Font(letra)
font_titulo = Font(size = 16, bold = True, color = '00FFFFFF')
font_subtitulo = Font(size = 14)
font_normal = Font(size = 11)


# 7.2.3.2 Otorgando los formatos a las celdas: Fill , Font , Alignment , Sides
# 7.2.3.2.1 Fill
ws[celda_titulo_general].fill = negro
#        Celdas de información
ws[celda_fecha_label].fill = gris
ws[celda_fecha ].fill = gris
ws[celda_documentos_label].fill = gris
ws[celda_documentos ].fill = gris
ws[celda_materiales_label].fill = gris
ws[celda_materiales ].fill = gris
ws[celda_materiales_con_matricula_label ].fill = gris
ws[celda_materiales_con_matricula].fill = gris
ws[celda_materiales_sin_matricula_label ].fill = gris
ws[celda_materiales_sin_matricula].fill = gris
#       Celdas de buscador
ws[celda_buscador_label ].fill = naranja_fuerte
ws[celda_matricula_label ].fill = naranja_debil 
ws[celda_matricula ].fill = naranja_debil
ws[celda_nombre_label ].fill = naranja_debil
ws[celda_nombre ].fill = naranja_debil
ws[celda_unidad_label ].fill = naranja_debil
ws[celda_unidad ].fill = naranja_debil
ws[celda_pu_label].fill = naranja_debil
ws[celda_pu ].fill = naranja_debil
ws[celda_cantidad_proveedor_label].fill = naranja_debil
ws[celda_cantidad_proveedor ].fill = naranja_debil
ws[celda_cantidad_contabilizada_label].fill = naranja_debil
ws[celda_cantidad_contabilizada ].fill = naranja_debil
#       Celdas de listado
ws[celda_listado_de_materiales_label ].fill = azul_fuerte
ws[celda_matricula_label_2 ].fill = azul_debil
ws[celda_nombre_label_2 ].fill = azul_debil
ws[celda_unidad_label_2 ].fill = azul_debil
ws[celda_pu_label_2 ].fill = azul_debil
ws[celda_cantidad_proveedor_label_2 ].fill = azul_debil
ws[celda_cantidad_contabilizada_label_2 ].fill = azul_debil


# 7.2.3.2.2 Font
ws[celda_titulo_general].font = font_titulo
#        Celdas de información
ws[celda_fecha_label].font =  font_normal 
ws[celda_fecha ].font =  font_normal 
ws[celda_documentos_label].font =  font_normal 
ws[celda_documentos ].font =  font_normal 
ws[celda_materiales_label].font =  font_normal 
ws[celda_materiales ].font =  font_normal 
ws[celda_materiales_con_matricula_label ].font =  font_normal 
ws[celda_materiales_con_matricula].font =  font_normal 
ws[celda_materiales_sin_matricula_label ].font =  font_normal 
ws[celda_materiales_sin_matricula].font =  font_normal 
#       Celdas de buscador
ws[celda_buscador_label ].font = font_normal 
ws[celda_matricula_label ].font =   font_normal 
ws[celda_matricula ].font =  font_normal 
ws[celda_nombre_label ].font =  font_normal 
ws[celda_nombre ].font =  font_normal 
ws[celda_unidad_label ].font =  font_normal 
ws[celda_unidad ].font =  font_normal 
ws[celda_pu_label].font =  font_normal 
ws[celda_pu ].font =  font_normal 
ws[celda_cantidad_proveedor_label].font =  font_normal 
ws[celda_cantidad_proveedor ].font =  font_normal 
ws[celda_cantidad_contabilizada_label].font =  font_normal 
ws[celda_cantidad_contabilizada ].font =  font_normal 
#       Celdas de listado
ws[celda_listado_de_materiales_label ].font = font_normal 
ws[celda_matricula_label_2 ].font =  font_normal 
ws[celda_nombre_label_2 ].font =  font_normal 
ws[celda_unidad_label_2 ].font =  font_normal 
ws[celda_pu_label_2 ].font =  font_normal 
ws[celda_cantidad_proveedor_label_2 ].font =  font_normal 
ws[celda_cantidad_contabilizada_label_2 ].font =  font_normal 


# 7.2.3.2.3 Alignment
ws[celda_titulo_general] .alignment = centro 
#        Celdas de información
ws[celda_fecha_label] .alignment =   izquierda  
ws[celda_fecha ] .alignment =    derecha 
ws[celda_documentos_label] .alignment =    izquierda 
ws[celda_documentos ] .alignment =    derecha 
ws[celda_materiales_label] .alignment =    izquierda 
ws[celda_materiales ] .alignment =    derecha 
ws[celda_materiales_con_matricula_label ] .alignment =  izquierda   
ws[celda_materiales_con_matricula] .alignment =   derecha  
ws[celda_materiales_sin_matricula_label ] .alignment =    izquierda 
ws[celda_materiales_sin_matricula] .alignment =    derecha 
#       Celdas de buscador
ws[celda_buscador_label ] .alignment =  centro  
ws[celda_matricula_label ] .alignment =     centro 
ws[celda_matricula ] .alignment =   izquierda  
ws[celda_nombre_label ] .alignment =    centro 
ws[celda_nombre ] .alignment =    centro
ws[celda_unidad_label ] .alignment =    centro 
ws[celda_unidad ] .alignment =    centro
ws[celda_pu_label] .alignment =    centro 
ws[celda_pu ] .alignment =    centro
ws[celda_cantidad_proveedor_label] .alignment =  centro   
ws[celda_cantidad_proveedor ] .alignment =    centro
ws[celda_cantidad_contabilizada_label] .alignment =   centro  
ws[celda_cantidad_contabilizada ] .alignment =  centro  
#       Celdas de listado
ws[celda_listado_de_materiales_label ] .alignment =   centro 
ws[celda_matricula_label_2 ] .alignment =    centro 
ws[celda_nombre_label_2 ] .alignment =    centro 
ws[celda_unidad_label_2 ] .alignment =    centro 
ws[celda_pu_label_2 ] .alignment =    centro 
ws[celda_cantidad_proveedor_label_2 ] .alignment =   centro  
ws[celda_cantidad_contabilizada_label_2 ] .alignment =  centro   


# 7.2.3.2.3 Side
# Titulo general
[celda_izquierda, celda_derecha] = celdas_titulo_general.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
# Datos
#   Fecha
[celda_izquierda, celda_derecha] = celdas_fecha_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Documentos
[celda_izquierda, celda_derecha] = celdas_documentos_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Materiales
[celda_izquierda, celda_derecha] = celdas_materiales_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Materiales con matrícula
[celda_izquierda, celda_derecha] = celdas_materiales_con_matricula_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Materiales sin matrícula
[celda_izquierda, celda_derecha] = celdas_materiales_sin_matricula_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Buscador
[celda_izquierda, celda_derecha] = celdas_buscador_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Matricula label
[celda_izquierda, celda_derecha] = celdas_matricula_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Matricula 
[celda_izquierda, celda_derecha] = celdas_matricula.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Nombre label 
[celda_izquierda, celda_derecha] = celdas_nombre_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Nombre  
[celda_izquierda, celda_derecha] = celdas_nombre.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   PU label  
[celda_izquierda, celda_derecha] = celdas_pu_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   PU 
[celda_izquierda, celda_derecha] = celdas_pu.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad proveedor label 
[celda_izquierda, celda_derecha] = celdas_cantidad_proveedor_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad proveedor 
[celda_izquierda, celda_derecha] = celdas_cantidad_proveedor.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad contabilizada label 
[celda_izquierda, celda_derecha] = celdas_cantidad_contabilizada_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad contabilizada
[celda_izquierda, celda_derecha] = celdas_cantidad_contabilizada.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)

# Celdas listado
#   celdas listado label
[celda_izquierda, celda_derecha] = celdas_listado_de_materiales_label.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Matricula label 2
[celda_izquierda, celda_derecha] = celdas_matricula_label_2.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   Nombre label 2
[celda_izquierda, celda_derecha] = celdas_nombre_label_2.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   PU label  2
[celda_izquierda, celda_derecha] = celdas_pu_label_2.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad proveedor label 2
[celda_izquierda, celda_derecha] = celdas_cantidad_proveedor_label_2.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)
#   cantidad contabilizada label 2
[celda_izquierda, celda_derecha] = celdas_cantidad_contabilizada_label_2.split(':')
bordes_en_combinacion(celda_izquierda , celda_derecha, borde)

# El resto de las celdas que no están combinadas
#   Se realizará mediante la función de bordes en combinación, sin embargo, se ingresará la misma
#   celda 2 veces. Supuestamente deberá funcionar
#   - Generamos una lista de celdas indivuduales y agregamos cada elemento con .append()
celdas_individuales = []
celdas_individuales.append(celda_fecha)
celdas_individuales.append(celda_documentos)
celdas_individuales.append(celda_materiales)
celdas_individuales.append(celda_materiales_con_matricula)
celdas_individuales.append(celda_materiales_sin_matricula)
celdas_individuales.append(celda_unidad_label)
celdas_individuales.append(celda_unidad)
celdas_individuales.append(celda_unidad_label_2)
# - Aplicamos los bordes
for celda in celdas_individuales:
    bordes_en_combinacion(celda , celda, borde)







#  7.3 Vaciar todos los datos 
#  Aquí ya empezaremos a vaciar todos los datos que tenemos, esto significa:
#  7.3.1 Vaciando los títulos y subtítulos
ws[celda_titulo_general] = titulo_general 
ws[celda_buscador_label ]  = buscador_titulo 
ws[celda_listado_de_materiales_label ]  = listado_titulo 
    
#  7.3.2 Vaciando los datos (información)
#  7.3.2.1 labels
ws[celda_fecha_label] = fecha_label 
ws[celda_documentos_label]  =  documentos_label 
ws[celda_materiales_label]  =  materiales_label 
ws[celda_materiales_con_matricula_label ]  =  materiales_con_matricula_label 
ws[celda_materiales_sin_matricula_label ]  = materiales_sin_matricula_label  
#  7.3.2.2 Info
ws[celda_fecha ] = fecha
ws[celda_documentos ]  =  cantidad_documentos_leidos
ws[celda_materiales ]  =  cantidad_materiales
ws[celda_materiales_con_matricula]  = cantidad_materiales_con_matricula
ws[celda_materiales_sin_matricula]  = cantidad_materiales_sin_matricula

#  7.3.3 Vaciando los labels de las tabals
#       Celdas de buscador
ws[celda_matricula_label ]  =   matricula_label 
ws[celda_nombre_label ]  =  nombre_label 
ws[celda_unidad_label ]  =  unidad_label 
ws[celda_pu_label]  =  precio_unitario 
ws[celda_cantidad_proveedor_label]  =  cantidad_proveedor 
ws[celda_cantidad_contabilizada_label]  =  cantidad_contabilizada    
#       Celdas de listado   
ws[celda_matricula_label_2 ] = matricula_label   
ws[celda_nombre_label_2 ] = nombre_label    
ws[celda_unidad_label_2 ] = unidad_label  
ws[celda_pu_label_2 ] = precio_unitario
ws[celda_cantidad_proveedor_label_2 ] = cantidad_proveedor    
ws[celda_cantidad_contabilizada_label_2 ] = cantidad_contabilizada     


#   7.3.4 Vaciando los datos de los diccionarios
#   7.3.4.1 Diccionario de materiales con matrícula completo
for material in diccionario_materiales_compelto.values():
    # 7.3.4.1.1 Instanciamos el renglón que usaremos, ya sea el primero o el siguiente
    reporte_doc.siguiente_renglon_()
    # 7.3.4.1.2 Se combinan las celdas necesarias para el listado
    # 7.3.4.1.2.1 Obteniendo las celdas necesarias
    combinacion_matricula = reporte_doc.celda_matricula_actual + ':' + 'B' + str( reporte_doc.renglon_actual )
    combinacion_nombre = reporte_doc.celda_nombre_actual + ':' + 'F' + str( reporte_doc.renglon_actual )
    combinacion_pu = reporte_doc.celda_pu_actual + ':' + 'I' + str( reporte_doc.renglon_actual )
    combinacion_cantidad_proveedor = reporte_doc.celda_cantidad_actual  + ':' + 'K' + str( reporte_doc.renglon_actual )
    combinacion_cantidad_contabilizada = reporte_doc.celda_cantidad_2_actual  + ':' + 'M' + str( reporte_doc.renglon_actual )
    # 7.3.4.1.2.2 Combinándolas
    ws.merge_cells(combinacion_matricula)
    ws.merge_cells(combinacion_nombre)
    ws.merge_cells(combinacion_pu)
    ws.merge_cells(combinacion_cantidad_proveedor)
    ws.merge_cells(combinacion_cantidad_contabilizada)
    
    
    # 7.3.4.1.3 Se les da color a las celdas principales respectivas a las combinaciones
    ws[ reporte_doc.celda_matricula_actual ].fill = azul_debil
    ws[ reporte_doc.celda_nombre_actual ].fill = azul_debil
    ws[ reporte_doc.celda_unidad_actual ].fill = azul_debil
    ws[ reporte_doc.celda_pu_actual ].fill = azul_debil
    ws[ reporte_doc.celda_cantidad_actual ].fill = azul_debil
    ws[ reporte_doc.celda_cantidad_2_actual ].fill = azul_debil
    
    
    # 7.3.4.1.4 Se les da color a las celdas principales respectivas a las combinaciones
    ws[ reporte_doc.celda_matricula_actual ].alignment = izquierda
    ws[ reporte_doc.celda_nombre_actual ].alignment = izquierda
    ws[ reporte_doc.celda_unidad_actual ].alignment = centro
    ws[ reporte_doc.celda_pu_actual ].alignment = centro
    ws[ reporte_doc.celda_cantidad_actual ].alignment = centro
    ws[ reporte_doc.celda_cantidad_2_actual ].alignment = centro
    
    
    # 7.3.4.1.5 Se ingresan los datos a las celdas
    ws[reporte_doc.celda_matricula_actual] = material.matricula
    ws[reporte_doc.celda_nombre_actual] = material.nombre
    ws[reporte_doc.celda_unidad_actual] = material.unidad
    ws[reporte_doc.celda_pu_actual] = material.precio_unitario
    # Para realizar el redondeo, primero requerimos que los datos sean strings o números, no None type
    if material.cantidad_proveedor == None:
        # Teoricamente no se imprime nada y el lo mismo que un Pass
        ws[reporte_doc.celda_cantidad_actual] = material.cantidad_proveedor
    else:
        ws[reporte_doc.celda_cantidad_actual] = str( round( float(material.cantidad_proveedor) , 3) )

    if material.cantidad_contabilizada == None:
        # Teoricamente no se imprime nada y el lo mismo que un Pass
        ws[reporte_doc.celda_cantidad_actual] = material.cantidad_contabilizada
    else:
        ws[reporte_doc.celda_cantidad_2_actual] = str( round( float(material.cantidad_contabilizada) , 3) )
        
        
        
# Materiales sin matrícula    
for material in diccionario_materiales_sin_matricula.values():
    # 7.3.4.1.1 Instanciamos el renglón que usaremos, ya sea el primero o el siguiente
    reporte_doc.siguiente_renglon_()
    # 7.3.4.1.2 Se combinan las celdas necesarias para el listado
    # 7.3.4.1.2.1 Obteniendo las celdas necesarias
    combinacion_matricula = reporte_doc.celda_matricula_actual + ':' + 'B' + str( reporte_doc.renglon_actual )
    combinacion_nombre = reporte_doc.celda_nombre_actual + ':' + 'F' + str( reporte_doc.renglon_actual )
    combinacion_pu = reporte_doc.celda_pu_actual + ':' + 'I' + str( reporte_doc.renglon_actual )
    combinacion_cantidad_proveedor = reporte_doc.celda_cantidad_actual  + ':' + 'K' + str( reporte_doc.renglon_actual )
    combinacion_cantidad_contabilizada = reporte_doc.celda_cantidad_2_actual  + ':' + 'M' + str( reporte_doc.renglon_actual )
    # 7.3.4.1.2.2 Combinándolas
    ws.merge_cells(combinacion_matricula)
    ws.merge_cells(combinacion_nombre)
    ws.merge_cells(combinacion_pu)
    ws.merge_cells(combinacion_cantidad_proveedor)
    ws.merge_cells(combinacion_cantidad_contabilizada)
    
    
    # 7.3.4.1.3 Se les da color a las celdas principales respectivas a las combinaciones
    ws[ reporte_doc.celda_matricula_actual ].fill = rojo
    ws[ reporte_doc.celda_nombre_actual ].fill = azul_debil
    ws[ reporte_doc.celda_unidad_actual ].fill = azul_debil
    ws[ reporte_doc.celda_pu_actual ].fill = azul_debil
    ws[ reporte_doc.celda_cantidad_actual ].fill = azul_debil
    ws[ reporte_doc.celda_cantidad_2_actual ].fill = azul_debil
    
    
    # 7.3.4.1.4 Se les da color a las celdas principales respectivas a las combinaciones
    ws[ reporte_doc.celda_matricula_actual ].alignment = izquierda
    ws[ reporte_doc.celda_nombre_actual ].alignment = izquierda
    ws[ reporte_doc.celda_unidad_actual ].alignment = centro
    ws[ reporte_doc.celda_pu_actual ].alignment = centro
    ws[ reporte_doc.celda_cantidad_actual ].alignment = centro
    ws[ reporte_doc.celda_cantidad_2_actual ].alignment = centro
    
    
    # 7.3.4.1.5 Se ingresan los datos a las celdas
    ws[reporte_doc.celda_matricula_actual] = material.matricula
    ws[reporte_doc.celda_nombre_actual] = material.nombre
    ws[reporte_doc.celda_unidad_actual] = material.unidad
    ws[reporte_doc.celda_pu_actual] = material.precio_unitario
    # Para realizar el redondeo, primero requerimos que los datos sean strings o números, no None type
    if material.cantidad_proveedor == None:
        # Teoricamente no se imprime nada y el lo mismo que un Pass
        ws[reporte_doc.celda_cantidad_actual] = material.cantidad_proveedor
    else:
        ws[reporte_doc.celda_cantidad_actual] = str( round( float(material.cantidad_proveedor) , 3) )

    if material.cantidad_contabilizada == None:
        # Teoricamente no se imprime nada y el lo mismo que un Pass
        ws[reporte_doc.celda_cantidad_actual] = material.cantidad_contabilizada
    else:
        ws[reporte_doc.celda_cantidad_2_actual] = str( round( float(material.cantidad_contabilizada) , 3) )


#  7.3.5 Vaciando las fórmulas del buscador
#  7.3.5.1 Generando las fórmulas del buscador
buscar_cmd = 'Lookup('
buscar_cmd = '=' + buscar_cmd
cierre = ')'

celda_matricula = celda_matricula
celda_inicio_matriculas = reporte_doc.columna_matricula + str(reporte_doc.renglon_inicio)
celda_fin_matriculas =  reporte_doc.columna_matricula + str(reporte_doc.renglon_actual )
rango_matriculas =  celda_inicio_matriculas + ':' + celda_fin_matriculas

celda_inicio_nombres =reporte_doc.columna_nombre + str(reporte_doc.renglon_inicio)
celda_fin_nombres =  reporte_doc.columna_nombre + str(reporte_doc.renglon_actual )
rango_nombres = celda_inicio_nombres + ':' + celda_fin_nombres

celda_inicio_unidades = reporte_doc.columna_unidad + str(reporte_doc.renglon_inicio)
celda_fin_unidades =  reporte_doc.columna_unidad + str(reporte_doc.renglon_actual )
rango_unidades = celda_inicio_unidades + ':' + celda_fin_unidades

celda_inicio_pu = reporte_doc.columna_pu + str(reporte_doc.renglon_inicio)
celda_fin_pu =  reporte_doc.columna_pu + str(reporte_doc.renglon_actual )
rango_pu = celda_inicio_pu + ':' + celda_fin_pu

celda_inicio_cantidad_proveedor = reporte_doc.columna_cantidad + str(reporte_doc.renglon_inicio)
celda_fin_cantidad_proveedor =  reporte_doc.columna_cantidad + str(reporte_doc.renglon_actual )
rango_cantidad_proveedor = celda_inicio_cantidad_proveedor + ':' + celda_fin_cantidad_proveedor

celda_inicio_cantidad_contabilizada = reporte_doc.columna_cantidad_2 + str(reporte_doc.renglon_inicio)
celda_fin_cantidad_contabilizada =  reporte_doc.columna_cantidad_2 + str(reporte_doc.renglon_actual )
rango_cantidad_contabilizada = celda_inicio_cantidad_contabilizada + ':' + celda_fin_cantidad_contabilizada


formula_inicio_default = buscar_cmd + celda_matricula + ',' + rango_matriculas + ','

formula_nombre = formula_inicio_default + rango_nombres + cierre
formula_unidad = formula_inicio_default + rango_unidades + cierre
formula_pu = formula_inicio_default + rango_pu + cierre
formula_cantidad_proveedor = formula_inicio_default + rango_cantidad_proveedor + cierre
formula_cantidad_contabilizada = formula_inicio_default + rango_cantidad_contabilizada + cierre
    
#  7.3.5.2 Vaciando fórmulas
ws[ celda_nombre ] = formula_nombre
ws[ celda_unidad ] = formula_unidad
ws[ celda_pu ] = formula_pu
ws[ celda_cantidad_proveedor ] = formula_cantidad_proveedor
ws[ celda_cantidad_contabilizada ] = formula_cantidad_contabilizada


#  7.4 Guardar el documento
fecha = datetime.now().strftime('%d_%b_%Y')
extension = '.xlsx'
nombre_final_reporte = directorio_destino + '\\' + nombre_reporte + ' ' + fecha + extension

wb.save(nombre_final_reporte)
print('...Documento guardado...')











